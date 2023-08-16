import os
import qrcode
from PIL import Image, ImageDraw, ImageFont
import random
import string
from openpyxl import Workbook
from datetime import datetime
import requests
import time
import json
from tqdm import tqdm

list = []
base_url = 'https://spotevents.co/'
#base_url = "http://localhost:3000/"

# to get the time


def get_time():
    day = datetime.now().day
    month = datetime.now().month
    hour = datetime.now().hour
    minute = datetime.now().minute
    if hour > 12:
        hour = hour - 12
        am_pm = "PM"
    elif hour == 0:
        hour = 12
        am_pm = "AM"
    else:
        am_pm = "AM"
    return f"[{day}-{month}] [{hour}~{minute} {am_pm}]"


dt_string = get_time()
wb = Workbook()
ws = wb.active

# to get permission to the system


def permission(tickets, type, server_sent):
    print("Please wait until we get the permission to generate the tickets")
    r = requests.post(
        f"{base_url}api/permission",
        json={
            "ticketNumber": tickets,
            "type": types[int(type) - 1],
            "server_sent": server_sent,
        },
    )
    return r.status_code


# for any error
def error(message):
    print("An error occurred because " + message)


# for image data
def chooseImage():
    currentDirectory = os.getcwd()
    files = os.listdir(currentDirectory)
    myFiles = []
    count = 1
    for f in files:
        if f.endswith(".jpg"):
            myFiles.append(f)
            print(f"{count}_ {f}")
            count += 1
    image = input("Enter the image name number: ")
    while image == "" or not image.isnumeric or int(image) > count:
        error("This image is not a valid image number")
        image = input("Enter the image name number: ")
    return myFiles[int(image) - 1]


# for code tickets data


def get_events():
    r = requests.get(f"{base_url}api/superadmin/1/events")
    return r.json()


def code_text(im, data, change_event):
    text_data = data
    ticket_needed = input("Do you want to add a text code (Y/N): ")
    while ticket_needed == "" or ticket_needed.upper() not in ["Y", "N"]:
        error("This is not a valid answer")
        ticket_needed = input("Do you want to add a text code (Y/N): ")
    if ticket_needed.upper() == "N":
        ticket_needed = False
        text_data["x"] = 0
        text_data["y"] = 0
        text_data["size"] = 0
        text_data["color"] = 0
        text_data["ticket_needed"] = ticket_needed
        return text_data
    if ticket_needed.upper() == "Y" and text_data != {} and change_event == False:
        text_data["ticket_needed"] = ticket_needed
        return text_data
    if ticket_needed.upper() == "Y" and  (change_event == True or text_data == {}):
        color_options = input("Do you want the color black or white (B/W): ")
        while color_options == "" or color_options.upper() not in ["B", "W"]:
            error("This is not a valid answer")
            color_options = input("Do you want the color black or white (B/W): ")
        if color_options.upper() == "W":
            text_data["color"] = (255, 255, 255)
        else:
            text_data["color"] = (0, 0, 0)
        ticket_needed = True
        text_data["x"] = input("Enter the text x-access: ")
        while (
            text_data["x"] == ""
            or text_data["x"].isnumeric() == False
            or int(text_data["x"]) > im.size[0]
            or int(text_data["x"]) < 0
        ):
            error("This x-access is not a valid number")
            text_data["x"] = input("Enter the text x-access: ")
        text_data["y"] = input("Enter the text y-access: ")
        while (
            text_data["y"] == ""
            or text_data["y"].isnumeric() == False
            or int(text_data["y"]) > im.size[1]
            or int(text_data["y"]) < 0
        ):
            error("This y-access is not a valid number")
            text_data["y"] = input("Enter the text y-access: ")
        text_data["size"] = input("Enter the text size: ")
        while (
            text_data["size"] == ""
            or text_data["size"].isnumeric() == False
            or int(text_data["size"]) > im.size[1]
            or int(text_data["size"]) > im.size[0]
            or int(text_data["size"]) < 0
        ):
            error("This text size is not a valid number")
            text_data["size"] = input("Enter the text size: ")
        text_data["ticket_needed"] = ticket_needed
        return text_data


# the serial number of the ticket
events = get_events()
print("Choose the event number: ")
for i in range(len(events)):
    print(f'{i+1}_ {events[i]["name"]}')
event = input("Enter the event number: ")
while event == "" or event.isnumeric() == False or int(event) > len(events):
    error("This event is not a valid event number")
    event = input("Enter the event number: ")
event = events[int(event) - 1]
print(f"you choosed {event['name']} event")
print("--------------------------------------------------")
print(f"qr data: {event['qr']}") if event["qr"] != None else print(
    "qr data: Empty"
)
print("text data: " + event["text"]) if event["text"] != None else print(
    "text data: Empty"
)
# ask user if he wants to change the event data
change_event = input("Do you want to change the event data (Y/N): ")
while change_event == "" or change_event.upper() not in ["Y", "N"]:
    error("This is not a valid answer")
    change_event = input("Do you want to change the event data (Y/N): ")
if change_event.upper() == "Y":
    change_event = True
else:
    change_event = False
qr_data = (
    json.loads(event["qr"]) if change_event == False and event["qr"] != None else {}
)
text_data = (
    json.loads(event["text"]) if change_event == False and event["text"] != None else {}
)
types = [key for key in json.loads(event["price"]).keys()]


im = chooseImage()
img_for_size = Image.open(im).convert("RGB")

# for ticket type data
for i in range(len(types)):
    print(f"{i+1}_ {types[i]}")
type = input("Enter the type of this codes: ")
while type == "" or type == type.isnumeric() == False or int(type) > len(types):
    error("This type is not a valid type")
    type = input("Enter the type of this codes: ")

# for QR code data
isQRcode = input("Do you need a QR code? (Y/N): ")
while isQRcode == "" or isQRcode.upper() not in ["Y", "N"]:
    error("This is not a valid answer plz type Y for yes or N for no")
    isQRcode = input("Do you need a QR code? (Y/N): ")
if isQRcode.upper() == "Y" and (change_event == True or qr_data == {}):
    qr_data["x"] = input("Enter the qrcode x-access: ")
    while qr_data["x"] == "" or qr_data["x"].isnumeric() == False:
        error("This x access is not a valid number")
        qr_data["x"] = input("Enter the qrcode x-access: ")
    qr_data["y"] = input("Enter the qrcode y-access: ")
    while qr_data["y"] == "" or qr_data["y"].isnumeric() == False:
        error("This y-access is not a valid number")
        qr_data["y"] = input("Enter the qrcode y-access: ")
    qr_data["border"] = input("Enter the border size of QR-code: ")
    if qr_data["border"] == "" or qr_data["border"].isnumeric() == False:
        qr_data["border"] = 0
    else:
        qr_data["border"] = int(qr_data["border"])
    qr_data["size"] = input("Enter the size of QR-code: ")
    if (
        qr_data["size"] == ""
        or qr_data["size"].isnumeric() == False
        or int(qr_data["size"]) < 1
    ):
        qr_data["size"] = 51
    else:
        qr_data["size"] = int(qr_data["size"])
text = code_text(img_for_size, text_data, change_event)
# for the number of tickets
tickets = input("Enter the number of tickets to generate: ")
while tickets == "" or tickets.isnumeric() == False or tickets == 0:
    error("This number of tickets is not a valid number")
    tickets = input("Enter the number of tickets to generate: ")

# for sent the data to the server
server_sent = input("Do you want to send tickets to the server (Y/N): ")
while server_sent == "" or server_sent.upper() not in ["Y", "N"]:
    error("This is not a valid answer")
    server_sent = input("Do you want to add a text code (Y/N): ")

# check perrmission to the tickets system
if permission(tickets, type, server_sent) != 200:
    error("You are not allowed to access this program at this moment")
    input("Press enter to exit")
    quit()


# for the random code
def random_password():
    pas = "".join(
        random.choice(string.ascii_lowercase + string.digits) for _ in range(int(12))
    )
    return pas[0:4] + "-" + pas[4:8] + "-" + pas[8:12]


# for genrating
def generate():
    for i in tqdm(
        range(int(tickets)), desc="Generating tickets… ", ascii=False, ncols=75
    ):
        pas = random_password()
        src = Image.open(im).convert("RGB")
        if isQRcode.upper() == "Y":
            qr = qrcode.make(pas, box_size=qr_data["size"], border=qr_data["border"])
            img = qr
            src.paste(img, (int(qr_data["x"]), int(qr_data["y"])))
        # for for the code_text
        if text["ticket_needed"] == True:
            draw = ImageDraw.Draw(src)
            draw.text(
                (int(text["x"]), int(text["y"])),
                f"{pas}",
                font=fontcode,
                fill=text["color"],
            )  # code
        src.save(f"{dt_string}/{i+1}.jpg")
        list.append(
            [
                pas,
                types[int(type) - 1],
                json.loads(event["price"])[types[int(type) - 1]],
                event["id"],
            ]
        )
        ws["A" + str(i + 2)].value = i + 1
        ws["B" + str(i + 2)].value = pas
        ws["C" + str(i + 2)].value = types[int(type) - 1]
        time.sleep(0.01)


# for saving the data
def create_folder(folder_name):
    currentDirectory = os.getcwd()
    if not os.path.exists(folder_name):
        os.mkdir(currentDirectory + f"\\{folder_name}")
        generate()
    else:
        print(
            f"folder[{folder_name}] is already exists please change the file name manually"
        )


fontcode = ImageFont.truetype("arial.ttf", size=int(text["size"]))

ws["A" + str(1)].value = "TicketNmber"
ws["B" + str(1)].value = "code"
ws["C" + str(1)].value = "Type"

create_folder(dt_string)
wb.save(f"{dt_string}/{dt_string}.xlsx")


if server_sent.upper() == "N":
    print("---------------- i'm done here ----------------")
else:
    print("---------------- i'm sending the data to the server ----------------")
    # ask user if he wants to update the event data
    update_event = input("Do you want to update the event data (Y/N): ")
    while update_event == "" or update_event.upper() not in ["Y", "N"]:
        error("This is not a valid answer")
        update_event = input("Do you want to update the event data (Y/N): ")
    if update_event.upper() == "Y":
        if "ticket_needed" in text_data:
            del text_data["ticket_needed"]
        text_data["color"] = "w" if text_data["color"] == [255, 255, 255] else "b"
        payload = {
            "text": text_data,
            "qr": qr_data,
        }
        # add header to the request
        headers = {
            "Content-Type": "application/json",
        }
        update = requests.post(
            f"{base_url}api/event/update/{event['id']}", json=payload, headers=headers
        )
        try:
            print(update.json())
        except Exception:
            print("error has been happend")
            print(update.text)

    r = requests.post(f"{base_url}api/insert_tickets", json={"code": list})
    try:
        print(r.json())
    except Exception:
        print("error has been happend")
        print(r.text)
    print("---------------- i'm done here ----------------")
k = input("Press Enter to exit")

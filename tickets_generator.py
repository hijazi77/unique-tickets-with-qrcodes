import os
import qrcode
from PIL import Image, ImageDraw, ImageFont
import random
import string
from openpyxl import Workbook
from datetime import datetime
import requests
import time
from tqdm import tqdm

list = []
types = ["Parent", "Student", "Guest", "test"]

# to get the time


def get_time():
    day = datetime.now().day
    month = datetime.now().month
    hour = datetime.now().hour
    minute = datetime.now().minute
    if hour > 12:
        hour = hour - 12
        am_pm = "PM"
    elif hour == 0:
        hour = 12
        am_pm = "AM"
    else:
        am_pm = "AM"
    return f"[{day}-{month}] [{hour}~{minute} {am_pm}]"


dt_string = get_time()
wb = Workbook()
ws = wb.active

# to get permission to the system


def permission(tickets, type):
    r = requests.post('https://api.obmeg.com/permission', json={
        "ticketNumber": tickets,
        "type": types[int(type)-1],
    })
    return r.status_code

# for any error


def error(message):
    print("An error occurred because "+message)

# for image data


def chooseImage():
    currentDirectory = os.getcwd()
    files = os.listdir(currentDirectory)
    myFiles = []
    count = 1
    for f in files:
        if f.endswith(".jpg"):
            myFiles.append(f)
            print(f'{count}_ {f}')
            count += 1
    image = input('Enter the image name number: ')
    while image == '' or not image.isnumeric or int(image) > count:
        error('This image is not a valid image number')
        image = input('Enter the image name number: ')
    return myFiles[int(image)-1]

# for code tickets data


def code_text(im):
    ticket_needed = input("Do you want to add a text code (Y/N): ")
    while ticket_needed == '' or ticket_needed.upper() not in ['Y', 'N']:
        error('This is not a valid answer')
        ticket_needed = input("Do you want to add a text code (Y/N): ")
    if ticket_needed.upper() == "N":
        ticket_needed = False
        x = 0
        y = 0
        ticket_size = 0
        color_options = 0
        return ticket_needed, x, y, ticket_size, color_options
    else:
        color_options = input("Do you want the color black or white (B/W): ")
        while color_options == '' or color_options.upper() not in ['B', 'W']:
            error('This is not a valid answer')
            color_options = input(
                "Do you want the color black or white (B/W): ")
        if color_options.upper() == "W":
            color_options = (255, 255, 255)
        else:
            color_options = (0, 0, 0)
        ticket_needed = True
        x = input('Enter the text x-access: ')
        while x == "" or x.isnumeric() == False or int(x) > im.size[0] or int(x) < 0:
            error('This x-access is not a valid number')
            x = input('Enter the text x-access: ')
        y = input('Enter the text y-access: ')
        while y == "" or y.isnumeric() == False or int(y) > im.size[1] or int(y) < 0:
            error('This y-access is not a valid number')
            y = input('Enter the text y-access: ')
        text_size = input('Enter the text size: ')
        while text_size == "" or text_size.isnumeric() == False or int(text_size) > im.size[1] or int(text_size) > im.size[0] or int(text_size) < 0:
            error('This text size is not a valid number')
            text_size = input('Enter the text size: ')
        return ticket_needed, x, y, text_size, color_options

# the serial number of the ticket


def counter_ticket(im):
    is_counter = input("Do you want to count the tickets (Y/N): ")
    while is_counter == '' or is_counter.upper() not in ['Y', 'N']:
        error('This is not a valid answer')
        is_counter = input("Do you want to count the tickets (Y/N): ")
    if is_counter.upper() == "Y":
        x = input('Enter the text x-access: ')
        while x == "" or x.isnumeric() == False or int(x) > im.size[0] or int(x) < 0:
            error('This x-access is not a valid number')
            x = input('Enter the text x-access: ')
        y = input('Enter the text y-access: ')
        while y == "" or y.isnumeric() == False or int(y) > im.size[1] or int(y) < 0:
            error('This y-access is not a valid number')
            y = input('Enter the text y-access: ')
        counter_size = input('Enter the text size: ')
        while counter_size == "" or counter_size.isnumeric() == False or int(counter_size) > im.size[1] or int(counter_size) > im.size[0] or int(counter_size) < 0:
            error('This text size is not a valid number')
            counter_size = input('Enter the text size: ')
        return {"x": x, "y": y, "size": counter_size}
    else:
        return False


im = chooseImage()
img_for_size = Image.open(im).convert("RGB")

# for ticket type data
for i in range(len(types)):
    print(f'{i+1}_ {types[i]}')
type = input("Enter the type of this codes: ")
while type == '' or type == type.isnumeric() == False or int(type) > len(types):
    error('This type is not a valid type')
    type = input("Enter the type of this codes: ")

# for QR code data
isQRcode = input("Do you need a QR code? (Y/N): ")
while isQRcode == "" or isQRcode.upper() not in ["Y", "N"]:
    error('This is not a valid answer plz type Y for yes or N for no')
    isQRcode = input("Do you need a QR code? (Y/N): ")
if isQRcode.upper() == "Y":
    x = input('Enter the qrcode x-access: ')
    while x == "" or x.isnumeric() == False:
        error('This x-access is not a valid number')
        x = input('Enter the qrcode x-access: ')

    y = input('Enter the qrcode y-access: ')
    while y == "" or y.isnumeric() == False:
        error('This y-access is not a valid number')
        y = input('Enter the qrcode y-access: ')

    border = input('Enter the border size of QR-code: ')
    if border == "" or border.isnumeric() == False:
        border = 0
    else:
        border = int(border)

    qrSize = input('Enter the size of QR-code: ')
    if qrSize == "" or qrSize.isnumeric() == False or int(qrSize) < 1:
        qrSize = 51
    else:
        qrSize = int(qrSize)


text = code_text(img_for_size)
ticket_counter = counter_ticket(img_for_size)

# for the number of tickets
tickets = input('Enter the number of tickets to generate: ')
while tickets == "" or tickets.isnumeric() == False or tickets == 0:
    error('This number of tickets is not a valid number')
    tickets = input('Enter the number of tickets to generate: ')

# check perrmission to the tickets system
if permission(tickets, type) != 200:
    error('You are not allowed to access this program at this moment')
    input('Press enter to exit')
    quit()

# for the random code


def random_password():
    pas = ''.join(random.choice(string.ascii_lowercase +
                                string.digits) for _ in range(int(12)))
    return pas[0:4]+"-"+pas[4:8]+"-"+pas[8:12]

# for genrating


def generate():
    for i in tqdm(range(int(tickets)),
                  desc="Generating tickets… ",
                  ascii=False, ncols=75):
        pas = random_password()
        src = Image.open(im).convert('RGB')
        if isQRcode.upper() == "Y":
            qr = qrcode.make(pas, box_size=qrSize, border=border)
            img = qr
            src.paste(img, (int(x), int(y)))
        # for the ticket counter
        if ticket_counter != False:
            draw = ImageDraw.Draw(src)
            font = ImageFont.truetype(
                "arial.ttf", int(ticket_counter["size"]))
            draw.text(
                (int(ticket_counter["x"]), int(ticket_counter["y"])), f'{i+1}', font=font, fill=text[4])

        # for for the
        if text[0] == True:
            draw = ImageDraw.Draw(src)
            draw.text((int(text[1]), int(text[2])), f'{pas}', font=fontcode,
                      fill=text[4])  # code
        src.save(f'{dt_string}/{i+1}.pdf')
        list.append([pas, types[int(type)-1]])
        ws["A"+str(i+2)].value = i+1
        ws["B"+str(i+2)].value = pas
        ws["C"+str(i+2)].value = types[int(type)-1]
        time.sleep(0.01)

# for saving the data


def create_folder(folder_name):
    currentDirectory = os.getcwd()
    if not os.path.exists(folder_name):
        os.mkdir(currentDirectory+f'\\{folder_name}')
        generate()
    else:
        print(
            f'folder[{folder_name}] is already exists please change the file name manually')


fontcode = ImageFont.truetype('arial.ttf', size=int(text[3]))
# for sent the data to the server
server_sent = input("Do you want to send tickets to the server (Y/N): ")
while server_sent == '' or server_sent.upper() not in ['Y', 'N']:
    error('This is not a valid answer')
    server_sent = input("Do you want to add a text code (Y/N): ")

ws['A'+str(1)].value = 'TicketNmber'
ws['B'+str(1)].value = 'code'
ws['C'+str(1)].value = 'Type'

create_folder(dt_string)
wb.save(f'{dt_string}/{dt_string}.xlsx')


if server_sent.upper() == "N":
    print("---------------- i'm done here ----------------")
else:
    r = requests.post('https://api.obmeg.com/add', json={
        "code": list
    })
    print(r.json())
    print("---------------- i'm done here ----------------")
k = input("Press Enter to exit")
